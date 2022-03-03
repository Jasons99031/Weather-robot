import openpyxl
try:
    wb = openpyxl.load_workbook('test.xlsx')
except:
    wb = openpyxl.Workbook()

sheet = wb['Sheet']

sheet['b1']='456'
print(sheet.max_column)
print(sheet.max_row)

wb.save('test.xlsx')


'''
i=65+0
j=1+2

# 利用 Workbook 建立一個新的工作簿
wb = openpyxl.Workbook()

# 取得第一個工作表
sheet = wb.worksheets[0]

# 設定 sheet 工作表 A1 儲存格內容為 "Hello Python, Hello Excel."
sheet[chr(i)+str(j)] = 'Hello Python.'

# 儲存檔案
workbook.save('test.xlsx')
'''