import requests 
import json

import openpyxl #CVS
X=65 
Y=1
i=0
'''***************************************************************************************************************
請求檔案
'''
def Req(date):

    url='https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467110&stname=%25E9%2587%2591%25E9%2596%2580&datepicker=' 
    CSV_URL = url+date
    #print(CSV_URL)

    with requests.Session() as s: #下載HTML
        download = s.get(CSV_URL) 
        #print(download.text)
        content=download.text

    with open("test.txt",'w') as f:#將資料存在筆記本
        f.write(content)

'''***************************************************************************************************************
CVS整理資料(單份)
'''
def CVS_arrange(date):
    f=open("test.txt",'r')
    #row = open("test_2.text","w")  #砍掉多餘的檔案
    i=0
    j=0
    h=0
    while i<13650:
        i+=1
        T = f.readline()
        if i>364 and T!='\n': #要標題315 不要364
            
            T = T.split()
            T = str(T)

            
            if T=='[]' or T=='[\'</tr>\']' or T=='[\'<tr>\']' or T=='' or h>24:
                T=T
            else:
                j+=1
                
                if j%17!=1:
                    
                    T = T.lstrip("[\'<td>")             #去頭去尾
                    T = T.rstrip('&nbsp;</td>\']')
                    try:    #將資料轉成數字
                        T = float(T)
                    except ValueError:
                        T = str(T)

                    if  h<10: #在前面加上時間(小時)
                        t='0'+str(h)
                    
                    else:
                        t=str(h)
                    #if h<25:
                    #row.write("\n"+t+"\n\n")

                    cvs(T,t,date)
                    
                    #return T
                                    #測試
                    T = str(T)
                    #print(T)
                    #row.write('\''+T+'\'')
                    #row.write(',')

                else:
                    
                    h+=1
                #row.write('\n')
                    
    #row.close                
    f.close
'''***************************************************************************************************************
轉成CVS
'''
def cvs(v,h,date):
    global X,Y,i 

    # 利用 Workbook 建立一個新的工作簿
    if i:
        wb = openpyxl.load_workbook(date+'.xlsx')
        
    else:
        wb = openpyxl.Workbook()
    
    s_n = wb.sheetnames[0]
    sheet = wb[s_n]
    


    # 取得第一個工作表
    #sheet = wb.worksheets[0]
    print(chr(X)+str(Y))
    
    sheet[chr(X)+str(Y)]=v
    i+=1
    X+=1
    
    if not(i%16):
        sheet[chr(X)+str(Y)]=h
        Y+=1
        X=65
    if (Y>24):
        sheet['R'+str(Y-1)]=date
        i=0##
        Y=1##
        X=65#
    
    

    #wb.save('總檔.xlsx')
    wb.save(date+'.xlsx')

'''***************************************************************************************************************
Json整理
'''
#def Json_arrange(date):
    









'''***************************************************************************************************************
整合資料
'''
def fusion(file_name,date):
    #row 1~....
    #column A~...
        
    try:
        W_wb = openpyxl.load_workbook(file_name+'.xlsx')    
    except:
        W_wb = openpyxl.Workbook()
    W_sheet_name = W_wb.sheetnames[0] #陣列第一[0]的名字
    W_sheet = W_wb[W_sheet_name]

    W_rm = W_sheet.max_row


    R_wb = openpyxl.load_workbook(date +'.xlsx')
    R_sheet_name = R_wb.sheetnames[0] #陣列第一[0]的名字
    R_sheet = R_wb[R_sheet_name]
    
    R_c=R_sheet.min_column
    R_cM=R_sheet.max_column
    
    R_r=R_sheet.min_row
    R_rM=R_sheet.max_row

    while R_r<=R_rM:

        while R_c<=R_cM:
            
            W_sheet[chr(R_c+64)+str(W_rm+1)] = R_sheet[chr(R_c+64)+str(R_r)].value
            #print(R_sheet[chr(R_c+64)+str(R_r)].value)
            R_c+=1
            W_wb.save(file_name + '.xlsx')

        R_r+=1
        W_rm+=1
        R_c=R_sheet.min_column


    

'''***************************************************************************************************************
日期
'''
def all_date(yN,mN,dN,yM,mM,dM,mod): #mod=0,1,2,3
                                     # (0) 單存輸出時間       
                                     # (1) 請求檔案
                                     # (2) 整合資料
                                     # (3) (1) and (2)
    Y=1
    M=1
    D=1
    file_name=str(yN)+'-'+str(mN)+'-'+str(dN)+'~'+str(yM)+'-'+str(mM)+'-'+str(dM)

    while Y:
        
        
        if yM==yN:
            Y=0
            print('Y')
        else:
            Y=1
            mN=1
            yN+=1

        while mN<12 and M:

                
            if not(Y) and mM==mN:
                M=0
                print('M')
            else:
                M=1
                dN=1
                mN+=1

            
            
            if (mN<8 and mN%2) or (mN>7 and not(mN%2)):
                Day=31
            else:
                if mN!=2:
                    Day=30
                else:
                    if ( not(yN%4) and (yN%100)) or not(yN%400) :

                        Day=29
                    else:
                        Day=28
            while dN<Day and D:

                if not(M) and dM+1==dN:
                    D=0
                    print('D')
                else:
                    D=1  

                    if mN<10:
                        sm = '0'+str(mN)
                    else:
                        sm = str(mN) 
                    if dN<10:
                        sd = '0'+str(dN)
                    else:
                        sd = str(dN)

                    data = str(yN)+'-'+str(sm)+'-'+str(sd)
                                        #mod=0,1,2,3
                                        # (0) 單存輸出時間      
                                        # (1) 請求檔案 and (0)
                                        # (2) 整合資料 and (0)
                                        # (3) (1) and (2)
                    print(data)
                    if mod==1 or mod==3:
                        Req(data)
                        #CVS_arrange(data)
                    if mod==2 or mod==3:
                        
                        fusion(file_name,data)

                    dN+=1
                

                
                
            #print(data)



'''***************************************************************************************************************
輸出設定
'''
#mod=0,1,2,3
# (0) 單存輸出時間      
# (1) 請求檔案 and (0)
# (2) 整合資料 and (0)
# (3) (1) and (2)
#N=min
#M=max

yN=2021
mN=1
dN=3

yM=2021
mM=2
dM=3

mod=3
'''****************************************************************************************************************
'''
all_date(yN,mN,dN,yM,mM,dM,mod)
'''****************************************************************************************************************
'''





