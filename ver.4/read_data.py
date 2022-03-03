from openpyxl import Workbook,load_workbook
import json
from mods.date import date_Record
import threading
import os

for Y in range(2015,2022): #可用threading 每年並排運算
    print(Y)
    
    wb = Workbook()
    ws1 = wb.worksheets[0]
    ws1.title="all_w"


    date = date_Record(Y,1,1,Y,12,31)
    date.loop()

    os.rename("day.txt",str(Y) + ".txt")

    path = 'json_data//' 

    with open(str(Y)+".txt","r") as r:#讀取Y年的天數

        i = r.read()
        n = i.split(',')  #切出每一天
    for i in range(len(n)):
        
        with open(path+str(n[i])+'.json') as j: #並排每天
            data = json.load(j)
            if(i==0):
                keys = []
                for key in data.keys():
                    keys.append(key)
            for x in range(1,len(keys)):
                c = i*16+1+x
                ws1.cell(row=c, column=1,value=n[i])
                ws1.cell(row=c, column=2,value=keys[x])
                for w in range(len(data[keys[x]])):
                    ws1.cell(row=c, column=3+w,value=data[keys[x]][w])
            #     keys[K][H] #K:2~16 H:0~23
            

    print(Y)
    wb.save('WA_'+str(Y)+'.xlsx')







'''
class Req(threading.Thread):
    def __init__(self,dete): 
        threading.Thread.__init__(self) #要用threading funtion 要叫run
        self.date=dete
        url='https://e-service.cwb.gov.tw/HistoryDataQuery/DayDataController.do?command=viewMain&station=467110&stname=%25E9%2587%2591%25E9%2596%2580&datepicker=' 
        self.CSV_URL = url+self.date#金門某年某月某日
'''



'''
並行運算每一年
從1月1號開始到12月31號
看當前sheet中col最長的地方往下接##還是這可以並行運算?

threading [keys[0~len(keys)]] (row(每一小)一種一種數值慢慢加?)


'''


'''
for(15~xx)
    用date輸入20xx/1/1 20xx/12/31
open(20xx-m-d) as r;
    data = json.load(r)

workbook = openyxl.load_workbook('金門_20xx')

sheet = workbook.worksheets[1]

keys = []
for key in data.keys():   #看有什麼key
    keys.append(key)


x = sheet.max_column
for i in range(len(keys)):
    
    sheet['B'+ x ] = data[keys[i]
    for j 0~23
        sheet[ chr(67+j)+ x ] = data[keys[i]][j]
    x+=1

workbook.save('金門_20xx')
'''
'''
 fill(填充類)：顏色等
'''
'''
wb = Workbook()#開個新的sheet

ws1 = wb. create_sheet("Sheet1")
'''

'''
with open('2015-02-07.json') as r:
    data = json.load(r)


keys = []
#print(data)
for key in data.keys():
    keys.append(key)
#print(data.keys())
print(keys)
print(len(keys))

'''
'''
wb = load_workbook('金門_2020.xlsx')
ws1 = wb. create_sheet("這是天氣歐")
ws1[chr(66)+'2'] = '我是鬼吧'
wb.save('金門_2020.xlsx')
'''





















